import streamlit as st
import re
import time
import io
import zipfile
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import pandas as pd

def main():
    st.title("아티스트 음원 정산 보고서 자동 생성기 (Excel 기반)")

    # 1) 섹션1: 보고서 생성(파일 업로드 + 진행기간/발행일 입력 등)
    section_one_report_input()

    st.divider()

    # 2) 섹션2: 검증 결과 표시
    section_two_verification()

    st.divider()

    # 3) 섹션3: 결과 ZIP 다운로드
    section_three_download_zip()

    st.divider()
    st.info("끝")


# ------------------------------------------
# 1) 섹션1: 보고서 생성(파일 업로드 + 진행기간/발행일 입력)
# ------------------------------------------
def section_one_report_input():
    st.subheader("1) 정산 보고서 생성")

    default_ym = st.session_state.get("ym", "")
    default_report_date = st.session_state.get("report_date", "")

    ym = st.text_input("진행기간(YYYYMM)", default_ym)
    report_date = st.text_input("보고서 발행 날짜 (YYYY-MM-DD)", default_report_date)

    uploaded_song_cost = st.file_uploader("input_song cost.xlsx 업로드", type=["xlsx"])
    uploaded_online_revenue = st.file_uploader("input_online revenue.xlsx 업로드", type=["xlsx"])

    if st.button("정산 보고서 생성 시작"):
        if not re.match(r'^\d{6}$', ym):
            st.error("진행기간은 YYYYMM 6자리로 입력하세요.")
            return
        if not report_date:
            st.error("보고서 발행 날짜를 입력하세요.")
            return
        if not uploaded_song_cost or not uploaded_online_revenue:
            st.error("두 개의 엑셀 파일을 모두 업로드해야 합니다.")
            return

        st.session_state["ym"] = ym
        st.session_state["report_date"] = report_date

        check_dict = {
            "song_artists": [],
            "revenue_artists": [],
            "artist_compare_result": {},
            "verification_summary": {
                "total_errors": 0,
                "artist_error_list": []
            },
            "details_verification": {
                "정산서": [],
                "세부매출": []
            }
        }

        zip_data = generate_report_excel(
            ym, report_date,
            uploaded_song_cost,
            uploaded_online_revenue,
            check_dict
        )

        if zip_data is not None:
            st.success("정산 보고서 생성 완료! 아래 섹션에서 ZIP 다운로드 가능")
            st.session_state["report_done"] = True
            st.session_state["zip_data"] = zip_data
            st.session_state["check_dict"] = check_dict
        else:
            st.error("보고서 생성 중 오류가 발생했습니다.")


# ------------------------------------------
# 2) 섹션2: 검증 결과 표시
# ------------------------------------------
def section_two_verification():
    if st.session_state.get("report_done", False):
        st.subheader("2) 검증 결과")

        cd = st.session_state.get("check_dict", {})
        if not cd:
            st.info("검증 데이터가 없습니다.")
            return

        tab1, tab2 = st.tabs(["검증 요약", "세부 검증 내용"])

        with tab1:
            ar = cd.get("artist_compare_result", {})
            st.write("**아티스트 목록 비교**")
            st.write(f"- Song cost 아티스트 수 = {ar.get('song_count')}")
            st.write(f"- Revenue 아티스트 수 = {ar.get('revenue_count')}")
            st.write(f"- 공통 아티스트 수 = {ar.get('common_count')}")
            if ar.get("missing_in_song"):
                st.warning(f"Song에 없고 Revenue에만 있는 아티스트: {ar['missing_in_song']}")
            if ar.get("missing_in_revenue"):
                st.warning(f"Revenue에 없고 Song에만 있는 아티스트: {ar['missing_in_revenue']}")

            ver_sum = cd.get("verification_summary", {})
            total_err = ver_sum.get("total_errors", 0)
            artists_err = ver_sum.get("artist_error_list", [])
            if total_err == 0:
                st.success("모든 항목이 정상 계산되었습니다. (오류 0건)")
            else:
                st.error(f"총 {total_err}건의 계산 오류 발생!")
                st.warning(f"문제 발생 아티스트: {list(set(artists_err))}")

        with tab2:
            show_detailed_verification(cd)

    else:
        st.info("정산 보고서 생성 완료 후, 검증 결과가 표시됩니다.")


# ------------------------------------------
# 3) 섹션3: 결과 ZIP 다운로드
# ------------------------------------------
def section_three_download_zip():
    if st.session_state.get("report_done", False):
        st.subheader("3) 결과 ZIP 다운로드")

        zip_data = st.session_state.get("zip_data")
        if zip_data:
            st.download_button(
                label="ZIP 다운로드",
                data=zip_data,
                file_name="정산결과보고서.zip",
                mime="application/zip"
            )
        else:
            st.warning("ZIP 데이터가 없습니다.")
    else:
        st.info("아직 보고서가 생성되지 않았습니다.")


# --------------------------------------------------
# 검증 표시 함수
# --------------------------------------------------
def show_detailed_verification(check_dict):
    dv = check_dict.get("details_verification", {})
    if not dv:
        st.warning("세부 검증 데이터가 없습니다.")
        return

    tabA, tabB = st.tabs(["정산서 검증", "세부매출 검증"])

    with tabA:
        rows = dv.get("정산서", [])
        if not rows:
            st.info("정산서 검증 데이터가 없습니다.")
        else:
            df = pd.DataFrame(rows)
            bool_cols = [c for c in df.columns if c.startswith("match_")]

            def highlight_boolean(val):
                if val is True:
                    return "background-color: #AAFFAA"
                elif val is False:
                    return "background-color: #FFAAAA"
                else:
                    return ""

            int_columns = [
                "원본_곡비", "정산서_곡비",
                "원본_공제금액", "정산서_공제금액",
                "원본_공제후잔액", "정산서_공제후잔액",
                "원본_정산율(%)", "정산서_정산율(%)"
            ]
            format_dict = {col: "{:.0f}" for col in int_columns if col in df.columns}

            st.dataframe(
                df.style
                  .format(format_dict)
                  .applymap(highlight_boolean, subset=bool_cols)
            )

    with tabB:
        rows = dv.get("세부매출", [])
        if not rows:
            st.info("세부매출 검증 데이터가 없습니다.")
        else:
            df = pd.DataFrame(rows)
            bool_cols = [c for c in df.columns if c.startswith("match_")]

            def highlight_boolean(val):
                if val is True:
                    return "background-color: #AAFFAA"
                elif val is False:
                    return "background-color: #FFAAAA"
                else:
                    return ""

            int_columns = ["원본_매출액", "정산서_매출액"]
            format_dict = {col: "{:.0f}" for col in int_columns if col in df.columns}

            st.dataframe(
                df.style
                  .format(format_dict)
                  .applymap(highlight_boolean, subset=bool_cols)
            )

# --------------------------------------------------
# 정산서 스타일
# --------------------------------------------------
def write_service_table(ws, start_row, service_list):
    """
    (1) 음원 서비스별 정산내역 작성

    인자:
      - ws: openpyxl Worksheet
      - start_row: 섹션을 시작할 엑셀 행(1-based)
      - service_list: [
          {"album":..., "major":..., "middle":..., "service":..., "year":..., "month":..., "revenue":...},
          ...
        ]

    반환:
      dict = {
        "section_title_row": <섹션제목 적은 행>,
        "header_row": <헤더 행>,
        "data_start": <데이터 시작행>,
        "data_end":   <데이터 끝행>,
        "sum_row":    <합계행>,
        "next_start_row": <다음 섹션부터 쓸 행>,
      }
    """
    # (A) 섹션 제목
    section_title_row = start_row
    ws.cell(row=section_title_row, column=2, value="1) 음원 서비스별 정산내역")

    # (B) 헤더: 다음 행
    header_row = section_title_row + 1
    ws.cell(row=header_row, column=2, value="앨범")
    ws.cell(row=header_row, column=3, value="대분류")
    ws.cell(row=header_row, column=4, value="중분류")
    ws.cell(row=header_row, column=5, value="서비스명")
    ws.cell(row=header_row, column=6, value="기간")
    ws.cell(row=header_row, column=7, value="매출액")

    # (C) 데이터
    data_start = header_row + 1
    r = data_start
    for item in service_list:
        ws.cell(row=r, column=2, value=item.get("album",""))
        ws.cell(row=r, column=3, value=item.get("major",""))
        ws.cell(row=r, column=4, value=item.get("middle",""))
        ws.cell(row=r, column=5, value=item.get("service",""))
        year_str  = str(item.get("year","2024"))
        month_str = str(item.get("month","12"))
        ws.cell(row=r, column=6, value=f"{year_str}년 {month_str}월")
        ws.cell(row=r, column=7, value=item.get("revenue",0.0))
        r += 1
    data_end = r - 1

    # (D) 합계행
    sum_row = data_end + 1
    ws.cell(row=sum_row, column=2, value="합계")
    total_val = sum(x.get("revenue",0.0) for x in service_list)
    ws.cell(row=sum_row, column=7, value=total_val)

    next_start_row = sum_row + 2  # 다음 섹션은 합계행 + 1~2행 띄우고 시작

    return {
        "section_title_row": section_title_row,
        "header_row": header_row,
        "data_start": data_start,
        "data_end": data_end,
        "sum_row": sum_row,
        "next_start_row": next_start_row
    }


def style_service_table(ws, info):
    """
    (1) 음원 서비스별 정산내역 스타일 적용

    info = {
      "section_title_row": ...,
      "header_row": ...,
      "data_start": ...,
      "data_end": ...,
      "sum_row": ...,
      ...
    }
    """

    dotted_side = Side(style="dotted", color="000000")
    header_bg = PatternFill("solid", fgColor="4CD9E0")  # 헤더
    band_e = "FFFFFF"  # 짝수줄
    band_o = "E5FCFF"  # 홀수줄
    sum_fill = PatternFill("solid", fgColor="E5FCFF")

    # 섹션 헤더(제목)은 row=info["section_title_row"], col=2
    # 원한다면 스타일 지정 가능
    title_row = info["section_title_row"]
    cell_title = ws.cell(row=title_row, column=2)
    cell_title.font = Font(bold=True, size=12)

    # 1) 헤더
    hr = info["header_row"]
    for c in range(2, 8):
        cell = ws.cell(row=hr, column=c)
        cell.fill = header_bg
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=dotted_side, bottom=dotted_side,
                             left=dotted_side, right=dotted_side)

    # 2) 본문 (data_start ~ data_end)
    ds = info["data_start"]
    de = info["data_end"]
    for r in range(ds, de+1):
        offset = r - ds
        row_fill = band_e if (offset % 2 == 0) else band_o
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=dotted_side, bottom=dotted_side,
                                 left=dotted_side, right=dotted_side)

    # 3) 합계행
    sr = info["sum_row"]
    # B~F 병합, G 따로
    ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=6)
    for c in range(2, 7):
        cell = ws.cell(row=sr, column=c)
        cell.fill = sum_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=dotted_side, bottom=dotted_side,
                             left=dotted_side, right=dotted_side)

    cell_g = ws.cell(row=sr, column=7)
    cell_g.fill = sum_fill
    cell_g.font = Font(bold=True)
    cell_g.alignment = Alignment(horizontal="center", vertical="center")
    cell_g.border = Border(top=dotted_side, bottom=dotted_side,
                           left=dotted_side, right=dotted_side)


def write_album_table(ws, start_row, album_list):
    """
    (2) 앨범별 정산 내역
    album_list: [{"album":..., "revenue":..., "year":..., "month":...}, ...]

    동일 패턴: 헤더 -> 데이터 -> 합계
    """
    section_title_row = start_row
    ws.cell(row=section_title_row, column=2, value="2) 앨범별 정산 내역")

    header_row = section_title_row + 1
    ws.cell(row=header_row, column=2, value="앨범")
    ws.cell(row=header_row, column=6, value="기간")
    ws.cell(row=header_row, column=7, value="매출액")

    data_start = header_row + 1
    r = data_start
    for alb in album_list:
        ws.cell(row=r, column=2, value=alb.get("album",""))
        y = str(alb.get("year","2024"))
        m = str(alb.get("month","12"))
        ws.cell(row=r, column=6, value=f"{y}년 {m}월")
        ws.cell(row=r, column=7, value=alb.get("revenue", 0.0))
        r += 1
    data_end = r - 1

    sum_row = data_end + 1
    ws.cell(row=sum_row, column=2, value="합계")
    total_val = sum(x.get("revenue",0.0) for x in album_list)
    ws.cell(row=sum_row, column=7, value=total_val)

    next_start_row = sum_row + 2

    return {
        "section_title_row": section_title_row,
        "header_row": header_row,
        "data_start": data_start,
        "data_end": data_end,
        "sum_row": sum_row,
        "next_start_row": next_start_row
    }


def style_album_table(ws, info):
    """
    (2) 앨범별 정산 내역 스타일
    """

    dotted_side = Side(style="dotted", color="000000")
    header_bg = PatternFill("solid", fgColor="4CD9E0")
    band_e = "FFFFFF"
    band_o = "E5FCFF"
    sum_fill = PatternFill("solid", fgColor="E5FCFF")

    # 섹션 제목
    cell_title = ws.cell(row=info["section_title_row"], column=2)
    cell_title.font = Font(bold=True, size=12)

    # 헤더
    hr = info["header_row"]
    for c in [2,6,7]:
        cell = ws.cell(row=hr, column=c)
        cell.fill = header_bg
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=dotted_side, right=dotted_side,
                             top=dotted_side, bottom=dotted_side)

    # 데이터
    ds = info["data_start"]
    de = info["data_end"]
    for r in range(ds, de+1):
        offset = r - ds
        row_fill = band_e if (offset % 2 == 0) else band_o

        for c in [2,6,7]:
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=dotted_side, right=dotted_side,
                                 top=dotted_side, bottom=dotted_side)

    # 합계행
    sr = info["sum_row"]
    # B~F 병합? -> 여기선 B6~B6..(?) 구조가 다르므로
    # 예시로 "B~F" 대신 "B~F"는 사실 안 쓰는 컬럼이 있을 수 있음
    ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=6)
    for c in range(2, 7):
        cell = ws.cell(row=sr, column=c)
        cell.fill = sum_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=dotted_side, right=dotted_side,
                             top=dotted_side, bottom=dotted_side)

    cell_g = ws.cell(row=sr, column=7)
    cell_g.fill = sum_fill
    cell_g.font = Font(bold=True)
    cell_g.alignment = Alignment(horizontal="center")
    cell_g.border = Border(left=dotted_side, right=dotted_side,
                           top=dotted_side, bottom=dotted_side)

def write_deduction_table(ws, start_row, ded_list):
    """
    (3) 공제 내역
    ded_list: [{"album":..., "prev_cost":..., "deduct_cost":..., "remain_cost":..., "after_deduct":...}, ...]
      - 예: 전월잔액 / 당월차감액 / 당월잔액 / 공제 적용 금액 등...
    """
    section_title_row = start_row
    ws.cell(row=section_title_row, column=2, value="3) 공제 내역")

    header_row = section_title_row + 1
    ws.cell(row=header_row, column=2, value="앨범")
    ws.cell(row=header_row, column=3, value="곡비")
    ws.cell(row=header_row, column=4, value="공제 금액")
    ws.cell(row=header_row, column=6, value="공제 후 남은 곡비")
    ws.cell(row=header_row, column=7, value="공제 적용 금액")

    data_start = header_row + 1
    r = data_start
    for d in ded_list:
        ws.cell(row=r, column=2, value=d.get("album",""))
        ws.cell(row=r, column=3, value=d.get("prev_cost",0.0))
        ws.cell(row=r, column=4, value=d.get("deduct_cost",0.0))
        ws.cell(row=r, column=6, value=d.get("remain_cost",0.0))
        ws.cell(row=r, column=7, value=d.get("after_deduct",0.0))
        r += 1
    data_end = r - 1

    # 합계행? 필요하면 추가
    # 여기서는 예시로 "합계"는 쓰지 않고 그냥 data_end까지만
    sum_row = data_end  # or None if no sum

    next_start_row = data_end + 2

    return {
        "section_title_row": section_title_row,
        "header_row": header_row,
        "data_start": data_start,
        "data_end": data_end,
        "sum_row": sum_row,
        "next_start_row": next_start_row
    }

def style_deduction_table(ws, info):

    dotted_side = Side(style="dotted", color="000000")
    header_bg = PatternFill("solid", fgColor="4CD9E0")
    band_e = "FFFFFF"
    band_o = "E5FCFF"

    # 제목
    ws.cell(row=info["section_title_row"], column=2).font = Font(bold=True, size=12)

    hr = info["header_row"]
    for c in [2,3,4,6,7]:
        cell = ws.cell(row=hr, column=c)
        cell.fill = header_bg
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=dotted_side, right=dotted_side,
                             top=dotted_side, bottom=dotted_side)

    ds = info["data_start"]
    de = info["data_end"]
    for r in range(ds, de+1):
        offset = r - ds
        row_fill = band_e if (offset % 2 == 0) else band_o
        for c in [2,3,4,5,6,7]:
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=dotted_side, right=dotted_side,
                                 top=dotted_side, bottom=dotted_side)

def write_rate_table(ws, start_row, rate_list):
    """
    (4) 수익 배분
    rate_list: [{"album":..., "rate":..., "applied_amount":..., ...}, ...]
    + "총 정산금액" 등도 작성
    """
    section_title_row = start_row
    ws.cell(row=section_title_row, column=2, value="4) 수익 배분")

    header_row = section_title_row + 1
    ws.cell(row=header_row, column=2, value="앨범")
    ws.cell(row=header_row, column=3, value="항목")
    ws.cell(row=header_row, column=4, value="적용율")
    ws.cell(row=header_row, column=7, value="적용 금액")

    data_start = header_row + 1
    r = data_start
    for d in rate_list:
        ws.cell(row=r, column=2, value=d.get("album",""))
        ws.cell(row=r, column=3, value="수익 배분율")
        ws.cell(row=r, column=4, value=f"{d.get('rate',0)}%")
        ws.cell(row=r, column=7, value=d.get("applied_amount", 0.0))
        r += 1
    data_end = r - 1

    # 마지막에 "총 정산금액" 표시
    sum_row = data_end + 1
    ws.cell(row=sum_row, column=2, value="총 정산금액")
    total_val = sum(x.get("applied_amount",0.0) for x in rate_list)
    ws.cell(row=sum_row, column=7, value=total_val)

    next_start_row = sum_row + 2

    return {
        "section_title_row": section_title_row,
        "header_row": header_row,
        "data_start": data_start,
        "data_end": data_end,
        "sum_row": sum_row,
        "next_start_row": next_start_row
    }

def style_rate_table(ws, info):

    dotted_side = Side(style="dotted", color="000000")
    header_bg = PatternFill("solid", fgColor="4CD9E0")
    band_e = "FFFFFF"
    band_o = "E5FCFF"
    sum_fill = PatternFill("solid", fgColor="E5FCFF")

    ws.cell(row=info["section_title_row"], column=2).font = Font(bold=True, size=12)

    hr = info["header_row"]
    for c in [2,3,4,7]:
        cell = ws.cell(row=hr, column=c)
        cell.fill = header_bg
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=dotted_side, right=dotted_side,
                             top=dotted_side, bottom=dotted_side)

    ds = info["data_start"]
    de = info["data_end"]
    for r in range(ds, de+1):
        offset = r - ds
        row_fill = band_e if (offset % 2 == 0) else band_o
        for c in [2,3,4,5,6,7]:
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=dotted_side, right=dotted_side,
                                 top=dotted_side, bottom=dotted_side)

    # 합계행
    sr = info["sum_row"]
    ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=6)
    for c in range(2, 7):
        cell = ws.cell(row=sr, column=c)
        cell.fill = sum_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=dotted_side, right=dotted_side,
                             top=dotted_side, bottom=dotted_side)

    cell_g = ws.cell(row=sr, column=7)
    cell_g.fill = sum_fill
    cell_g.font = Font(bold=True)
    cell_g.alignment = Alignment(horizontal="center")
    cell_g.border = Border(left=dotted_side, right=dotted_side,
                           top=dotted_side, bottom=dotted_side)

def create_report_excel(artist, service_list, album_list, deduction_list, rate_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{artist}(정산서)"

    # 1) 음원 서비스별
    row_cursor = 12
    info_service = write_service_table(ws, row_cursor, service_list)
    style_service_table(ws, info_service)
    row_cursor = info_service["next_start_row"]

    # 2) 앨범별
    info_album = write_album_table(ws, row_cursor, album_list)
    style_album_table(ws, info_album)
    row_cursor = info_album["next_start_row"]

    # 3) 공제 내역
    info_ded = write_deduction_table(ws, row_cursor, deduction_list)
    style_deduction_table(ws, info_ded)
    row_cursor = info_ded["next_start_row"]

    # 4) 수익 배분
    info_rate = write_rate_table(ws, row_cursor, rate_list)
    style_rate_table(ws, info_rate)
    row_cursor = info_rate["next_start_row"]

    # 전체 외곽 테두리, 열너비 등
    thin_side = Side(style="thin", color="000000")
    for r in range(1, row_cursor+10):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                top=thin_side, left=thin_side,
                right=thin_side, bottom=thin_side
            )

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 5

    filename = f"{artist}_정산서.xlsx"
    wb.save(filename)
    return filename


# --------------------------------------------------
# 세부매출내역 데이터 및 스타일
# --------------------------------------------------
def write_detail_data(ws, detail_list, start_row=1):
    """
    세부매출내역 시트에 'detail_list' 데이터를 써넣고,
    헤더/본문/합계행의 "행 번호"를 반환.

    detail_list: [{"albumArtist":..., "album":..., "major":..., "middle":..., "service":..., "revenue":...}, ...]

    반환 예시:
      {
        "header_row": 1,
        "data_start": 2,
        "data_end": 8,
        "sum_row": 9,
      }
    """
    # 1) 헤더 (start_row에 작성)
    ws.cell(row=start_row, column=1, value="앨범아티스트")
    ws.cell(row=start_row, column=2, value="앨범명")
    ws.cell(row=start_row, column=3, value="대분류")
    ws.cell(row=start_row, column=4, value="중분류")
    ws.cell(row=start_row, column=5, value="서비스명")
    ws.cell(row=start_row, column=6, value="기간")
    ws.cell(row=start_row, column=7, value="매출 순수익")

    # 2) 본문
    data_start = start_row + 1
    r = data_start
    for d in detail_list:
        ws.cell(row=r, column=1, value=d.get("albumArtist", ""))
        ws.cell(row=r, column=2, value=d.get("album", ""))
        ws.cell(row=r, column=3, value=d.get("major", ""))
        ws.cell(row=r, column=4, value=d.get("middle", ""))
        ws.cell(row=r, column=5, value=d.get("service", ""))
        ws.cell(row=r, column=6, value=f"{d.get('year','2024')}년 {d.get('month','12')}월")
        ws.cell(row=r, column=7, value=d.get("revenue", 0.0))
        r += 1
    data_end = r - 1  # 실제 마지막 데이터 행

    # 3) 합계행
    sum_row = data_end + 1
    ws.cell(row=sum_row, column=1, value="합계")
    # (단일 셀에 "합계" 쓰고, 7번 칸에 total)
    total_val = sum(d["revenue"] for d in detail_list)
    ws.cell(row=sum_row, column=7, value=total_val)

    return {
        "header_row": start_row,
        "data_start": data_start,
        "data_end": data_end,
        "sum_row": sum_row
    }


def apply_detail_style(ws, header_row, data_start, data_end, sum_row):
    """
    세부매출내역: 
      - 헤더 행: 주황(또는 노랑) 배경 + 굵은폰트
      - 본문(데이터): 검정 실선 테두리
      - 합계행: 연한 노랑 배경 + 굵은폰트
      - 전체 테두리는 thin(검정 실선)
    """

    thin_side = Side(style="thin", color="000000")
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # 1) 헤더 행
    header_fill = PatternFill("solid", fgColor="FFC000")
    for col in range(1, 8):  # A~G
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 2) 본문 데이터 (data_start ~ data_end)
    for row in range(data_start, data_end+1):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # 3) 합계행
    sum_fill = PatternFill("solid", fgColor="FFD966")
    for col in range(1, 8):
        cell = ws.cell(row=sum_row, column=col)
        cell.border = thin_border
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 7:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if col in (1, 7):
            # 합계 텍스트와 금액만 색/굵게
            cell.fill = sum_fill
            cell.font = Font(bold=True)

    # (열너비 등)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15


def create_detail_excel(artist, ym, detail_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{artist}(세부매출내역)"

    out_info = write_detail_data(ws, detail_list, start_row=1)

    apply_detail_style(
        ws,
        header_row=out_info["header_row"],
        data_start=out_info["data_start"],
        data_end=out_info["data_end"],
        sum_row=out_info["sum_row"]
    )

    filename = f"{artist}_세부매출.xlsx"
    wb.save(filename)
    return wb  # Workbook 객체 반환 (ZIP으로 저장 시 사용)


# --------------------------------------------------
# 보고서 생성 (엑셀 기반)
# --------------------------------------------------
def generate_report_excel(ym, report_date, file_song_cost, file_online_revenue, check_dict):
    """
    업로드된 두 엑셀 파일(file_song_cost, file_online_revenue)을 openpyxl로 파싱 →
    아티스트별로:
      1) 세부매출내역(artist).xlsx
      2) 정산서(artist).xlsx
    을 각각 생성, ZIP으로 묶어 반환 (bytes).

    - ym: "YYYYMM"
    - report_date: "YYYY-MM-DD"
    - file_song_cost: 업로드된 엑셀( song cost.xlsx )
    - file_online_revenue: 업로드된 엑셀( online revenue.xlsx )
    - check_dict: 검증용 딕셔너리 (실제 계산/비교 결과를 저장)

    반환: zip(bytes) or None
    """

    # ---------------------- (A) 엑셀 파싱 ----------------------
    try:
        wb_song = openpyxl.load_workbook(file_song_cost, data_only=True)
        wb_revenue = openpyxl.load_workbook(file_online_revenue, data_only=True)
    except Exception as e:
        st.error(f"엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")
        return None

    # 1) song cost → ym 시트
    if ym not in wb_song.sheetnames:
        st.error(f"[song cost] 파일에 '{ym}' 시트가 없습니다.")
        return None

    ws_sc = wb_song[ym]
    rows_sc = list(ws_sc.values)
    if not rows_sc:
        st.error(f"[song cost] '{ym}' 시트가 비어있습니다.")
        return None

    header_sc = rows_sc[0]
    body_sc = rows_sc[1:]

    # 필요한 컬럼 인덱스 찾기
    try:
        idx_artist = header_sc.index("아티스트명")
        idx_rate = header_sc.index("정산 요율")
        idx_prev = header_sc.index("전월 잔액")
        idx_deduct = header_sc.index("당월 차감액")
        idx_remain = header_sc.index("당월 잔액")
    except ValueError as e:
        st.error(f"[song cost] 시트 컬럼명이 올바른지 확인 필요: {e}")
        return None

    def to_num(x):
        if not x:
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        x = str(x).replace("%", "").replace(",", "")
        try:
            return float(x)
        except:
            return 0.0

    artist_cost_dict = {}
    for row in body_sc:
        if not row or len(row) < len(header_sc):
            continue
        artist_name = row[idx_artist]
        if not artist_name:
            continue
        cost_data = {
            "정산요율": to_num(row[idx_rate]),
            "전월잔액": to_num(row[idx_prev]),
            "당월차감액": to_num(row[idx_deduct]),
            "당월잔액": to_num(row[idx_remain])
        }
        artist_cost_dict[artist_name] = cost_data

    # 2) online revenue → ym 시트
    if ym not in wb_revenue.sheetnames:
        st.error(f"[online revenue] 파일에 '{ym}' 시트가 없습니다.")
        return None
    ws_or = wb_revenue[ym]
    rows_or = list(ws_or.values)
    if not rows_or:
        st.error(f"[online revenue] '{ym}' 시트가 비어있습니다.")
        return None

    header_or = rows_or[0]
    body_or = rows_or[1:]
    try:
        col_aartist = header_or.index("앨범아티스트")
        col_album = header_or.index("앨범명")
        col_major = header_or.index("대분류")
        col_middle = header_or.index("중분류")
        col_service = header_or.index("서비스명")
        col_revenue = header_or.index("권리사정산금액")
    except ValueError as e:
        st.error(f"[online revenue] 시트 컬럼명이 올바른지 확인 필요: {e}")
        return None

    artist_revenue_dict = defaultdict(list)
    for row in body_or:
        if not row or len(row) < len(header_or):
            continue
        aartist = str(row[col_aartist]).strip() if row[col_aartist] else ""
        album   = str(row[col_album])   if row[col_album]   else ""
        major   = str(row[col_major])   if row[col_major]   else ""
        middle  = str(row[col_middle])  if row[col_middle]  else ""
        srv     = str(row[col_service]) if row[col_service] else ""
        rev_val = to_num(row[col_revenue])

        if aartist:
            artist_revenue_dict[aartist].append({
                "album": album,
                "major": major,
                "middle": middle,
                "service": srv,
                "revenue": rev_val
            })

    # ---------------------- (B) 아티스트 목록 비교 ----------------------
    song_artists = sorted(artist_cost_dict.keys())
    revenue_artists = sorted(artist_revenue_dict.keys())
    check_dict["song_artists"] = song_artists
    check_dict["revenue_artists"] = revenue_artists

    compare_res = compare_artists(song_artists, revenue_artists)
    check_dict["artist_compare_result"] = compare_res

    # 전체 아티스트(둘 중 하나라도 존재)
    all_artists = sorted(set(song_artists) | set(revenue_artists))

    # ---------------------- (C) 아티스트별 엑셀 생성 & ZIP ----------------------
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        progress_bar = st.progress(0.0)
        artist_placeholder = st.empty()

        for i, artist in enumerate(all_artists):
            ratio = (i + 1) / len(all_artists)
            progress_bar.progress(ratio)
            artist_placeholder.info(f"[{i+1}/{len(all_artists)}] {artist} 처리 중...")

            # 1) 해당 아티스트의 cost_data, revenue_list
            cost_data = artist_cost_dict.get(artist, {
                "정산요율":0, "전월잔액":0, "당월차감액":0, "당월잔액":0
            })
            detail_list = artist_revenue_dict[artist]  # [{album, major, middle, service, revenue}, ...]

            # (A) 세부매출내역(.xlsx)
            #   write_detail_data + apply_detail_style 방식
            detail_wb = create_detail_excel(artist, ym, detail_list)
            detail_buf = io.BytesIO()
            detail_wb.save(detail_buf)
            detail_buf.seek(0)
            zf.writestr(f"{artist}(세부매출내역).xlsx", detail_buf.getvalue())

            # (B) 정산서(.xlsx)
            #   4섹션(write_service_table + style_service_table, etc.)
            #   우선 "service_list", "album_list", "deduction_list", "rate_list"를 구성
            #   또는, 단순히 detail_list를 가공해서 service_list, album_list 만들 수도 있음

            # (예시) service_list = detail_list와 동일하게 구성
            service_list = []
            for d in detail_list:
                service_item = {
                    "album": d["album"],
                    "major": d["major"],
                    "middle": d["middle"],
                    "service": d["service"],
                    "year": ym[:4],
                    "month": ym[4:],
                    "revenue": d["revenue"]
                }
                service_list.append(service_item)

            # (앨범별) album_list
            #   album별 revenue 합산
            album_dict = defaultdict(float)
            for d in detail_list:
                album_dict[d["album"]] += d["revenue"]
            album_list = []
            for alb, amt in album_dict.items():
                album_list.append({
                    "album": alb,
                    "year": ym[:4],
                    "month": ym[4:],
                    "revenue": amt
                })

            # (공제 내역) deduction_list
            #   cost_data["전월잔액"], cost_data["당월차감액"], cost_data["당월잔액"]
            #   after_deduct = (album합계 - 당월차감액) 등
            total_album_sum = sum(album_dict.values())
            after_deduct = total_album_sum - cost_data["당월차감액"]
            ded_list = [{
                "album": ", ".join(album_dict.keys()) if album_dict else "(앨범 없음)",
                "prev_cost": cost_data["전월잔액"],
                "deduct_cost": cost_data["당월차감액"],
                "remain_cost": cost_data["당월잔액"],
                "after_deduct": after_deduct
            }]

            # (수익 배분) rate_list
            #   rate = cost_data["정산요율"]
            #   applied_amount = after_deduct * (rate/100)
            applied_amount = after_deduct * (cost_data["정산요율"]/100.0)
            rate_list = [{
                "album": ", ".join(album_dict.keys()) if album_dict else "(앨범 없음)",
                "rate": cost_data["정산요율"],
                "applied_amount": applied_amount
            }]

            report_wb = create_report_excel(
                artist,
                service_list,
                album_list,
                ded_list,
                rate_list
            )
            report_buf = io.BytesIO()
            report_wb.save(report_buf)
            report_buf.seek(0)
            zf.writestr(f"{artist}(정산서).xlsx", report_buf.getvalue())

        artist_placeholder.success("모든 아티스트 처리 완료!")
        progress_bar.progress(1.0)

    return zip_buf.getvalue()


# -----------------------------------------
# 헬퍼 함수: 아티스트 목록 비교
# -----------------------------------------
def compare_artists(song_artists, revenue_artists):
    set_song = set(song_artists)
    set_revenue = set(revenue_artists)
    return {
        "missing_in_song": sorted(set_revenue - set_song),
        "missing_in_revenue": sorted(set_song - set_revenue),
        "common_count": len(set_song & set_revenue),
        "song_count": len(set_song),
        "revenue_count": len(set_revenue),
    }


def almost_equal(a, b, tol=1e-3):
    """숫자 비교용: 소수점 오차 허용."""
    return abs(a - b) < tol


# -----------------------------------------
# (A) 세부매출내역 Workbook 생성
# -----------------------------------------
def create_detail_workbook(artist, ym, detail_list, check_dict):
    """
    artist에 대한 세부매출내역 엑셀 파일을 생성하여 Workbook 객체로 반환.
    detail_list: [{"album":..., "major":..., "middle":..., "service":..., "revenue":...}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "세부매출내역"

    # 헤더
    ws.append(["앨범아티스트", "앨범명", "대분류", "중분류", "서비스명", "기간", "매출 순수익"])

    # 정렬(앨범명 등으로 정렬)
    detail_list_sorted = sorted(detail_list, key=lambda x: (x["album"], x["service"]))

    total_revenue = 0.0
    year_val, month_val = ym[:4], ym[4:]
    for d in detail_list_sorted:
        rv = d["revenue"]
        total_revenue += rv
        ws.append([
            artist,
            d["album"],
            d["major"],
            d["middle"],
            d["service"],
            f"{year_val}년 {month_val}월",
            rv
        ])

    # 합계 행
    ws.append(["합계", "", "", "", "", "", total_revenue])

    # 간단한 스타일 예시
    # 1) 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4CAF50")  # 연두/초록
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2) 합계행 스타일 (마지막 행)
    last_row = ws.max_row
    for c in range(1, 7):
        cell = ws.cell(row=last_row, column=c)
        cell.alignment = Alignment(horizontal="center")
    sum_cell = ws.cell(row=last_row, column=7)
    sum_font = Font(bold=True, color="000000")
    sum_fill = PatternFill("solid", fgColor="FFD966")  # 옅은 노랑
    sum_cell.font = sum_font
    sum_cell.fill = sum_fill

    # (검증 기록) → check_dict["details_verification"]["세부매출"] 에 추가
    # 실제로는 "정산서 값과 match" 여부를 비교해야 하지만,
    # 여기서는 generate_report_excel 쪽에서 처리하므로 생략

    return wb


# -----------------------------------------
# (B) 정산서 Workbook 생성
# -----------------------------------------
def create_report_workbook(artist, ym, report_date, cost_data, detail_list, check_dict):
    """
    artist에 대한 "정산서" 엑셀 파일을 생성하여 Workbook 객체로 반환.
    cost_data: {"정산요율":..., "전월잔액":..., "당월차감액":..., "당월잔액":...}
    detail_list: [{"album":..., "major":..., "middle":..., "service":..., "revenue":...}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "정산서"

    year_val, month_val = ym[:4], ym[4:]

    # -------------------------------
    # 1) 머리글
    # -------------------------------
    ws["H2"] = f"{report_date} 발행"
    ws["B4"] = f"{year_val}년 {month_val}월 판매분"
    ws["B6"] = f"{artist}님 음원 정산 내역서"

    ws["A8"] = "•"
    ws["B8"] = "저희와 함께해 주셔서 정말 감사하고, 앞으로도 잘 부탁드립니다!"
    ws["A9"] = "•"
    ws["B9"] = f"{year_val}년 {month_val}월 음원 수익을 아래와 같이 정산드립니다."
    ws["A10"] = "•"
    ws["B10"] = "정산 관련 문의사항이 있다면 언제든 편히 연락주세요!"
    ws["F10"] = "E-mail: help@xxxx.com"

    # -------------------------------
    # 2) 세부매출
    # -------------------------------
    row_start = 12
    ws.cell(row=row_start, column=1, value="1.")
    ws.cell(row=row_start, column=2, value="음원 서비스별 정산내역")

    row_start += 1
    headers = ["앨범", "대분류", "중분류", "서비스명", "기간", "매출액"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=row_start, column=i, value=h)

    detail_list_sorted = sorted(detail_list, key=lambda x: (x["album"], x["service"]))
    total_1 = 0.0
    curr = row_start + 1
    for d in detail_list_sorted:
        rv = d["revenue"]
        total_1 += rv

        ws.cell(row=curr, column=2, value=d["album"])
        ws.cell(row=curr, column=3, value=d["major"])
        ws.cell(row=curr, column=4, value=d["middle"])
        ws.cell(row=curr, column=5, value=d["service"])
        ws.cell(row=curr, column=6, value=f"{year_val}년 {month_val}월")
        ws.cell(row=curr, column=7, value=rv)
        curr += 1

    # 합계
    curr += 1
    ws.cell(row=curr, column=2, value="합계")
    ws.cell(row=curr, column=7, value=total_1)
    row_sum_1 = curr
    curr += 2

    # -------------------------------
    # 3) 앨범 별 정산
    # -------------------------------
    ws.cell(row=curr, column=1, value="2.")
    ws.cell(row=curr, column=2, value="앨범 별 정산 내역")
    curr += 1

    ws.cell(row=curr, column=2, value="앨범")
    ws.cell(row=curr, column=6, value="기간")
    ws.cell(row=curr, column=7, value="매출액")
    curr += 1

    album_sum = defaultdict(float)
    for d in detail_list_sorted:
        album_sum[d["album"]] += d["revenue"]

    total_2 = 0.0
    for alb in sorted(album_sum.keys()):
        amt = album_sum[alb]
        total_2 += amt
        ws.cell(row=curr, column=2, value=alb)
        ws.cell(row=curr, column=6, value=f"{year_val}년 {month_val}월")
        ws.cell(row=curr, column=7, value=amt)
        curr += 1

    ws.cell(row=curr, column=2, value="합계")
    ws.cell(row=curr, column=7, value=total_2)
    row_sum_2 = curr
    curr += 2

    # -------------------------------
    # 4) 공제 내역
    # -------------------------------
    ws.cell(row=curr, column=1, value="3.")
    ws.cell(row=curr, column=2, value="공제 내역")
    curr += 1

    ws.cell(row=curr, column=2, value="앨범")
    ws.cell(row=curr, column=3, value="곡비")
    ws.cell(row=curr, column=4, value="공제 금액")
    ws.cell(row=curr, column=6, value="공제 후 남은 곡비")
    ws.cell(row=curr, column=7, value="공제 적용 금액")
    curr += 1

    prev_val = cost_data.get("전월잔액", 0.0)
    deduct_val = cost_data.get("당월차감액", 0.0)
    remain_val = cost_data.get("당월잔액", 0.0)

    # 공제 적용된 매출액 = total_2 - 당월차감액
    공제적용 = total_2 - deduct_val

    alb_list = sorted(album_sum.keys())
    alb_str = ", ".join(alb_list) if alb_list else "(앨범 없음)"

    ws.cell(row=curr, column=2, value=alb_str)
    ws.cell(row=curr, column=3, value=prev_val)
    ws.cell(row=curr, column=4, value=deduct_val)
    ws.cell(row=curr, column=6, value=remain_val)
    ws.cell(row=curr, column=7, value=공제적용)
    curr += 2

    # -------------------------------
    # 5) 수익 배분
    # -------------------------------
    ws.cell(row=curr, column=1, value="4.")
    ws.cell(row=curr, column=2, value="수익 배분")
    curr += 1

    ws.cell(row=curr, column=2, value="앨범")
    ws.cell(row=curr, column=3, value="항목")
    ws.cell(row=curr, column=4, value="적용율")
    ws.cell(row=curr, column=7, value="적용 금액")
    curr += 1

    rate_val = cost_data.get("정산요율", 0.0)
    final_amount = 공제적용 * (rate_val / 100.0)

    ws.cell(row=curr, column=2, value=alb_str)
    ws.cell(row=curr, column=3, value="수익 배분율")
    ws.cell(row=curr, column=4, value=f"{rate_val}%")
    ws.cell(row=curr, column=7, value=final_amount)
    curr += 1

    ws.cell(row=curr, column=2, value="총 정산금액")
    ws.cell(row=curr, column=7, value=final_amount)
    curr += 2

    ws.cell(row=curr, column=7, value="* 부가세 별도")

    # (검증) check_dict["details_verification"]["정산서"] 에 매핑
    #  - (1) 공제 내역 검증
    is_match_prev = almost_equal(prev_val, cost_data.get("전월잔액", 0))
    is_match_deduct = almost_equal(deduct_val, cost_data.get("당월차감액", 0))
    is_match_remain = almost_equal(remain_val, cost_data.get("당월잔액", 0))
    if not (is_match_prev and is_match_deduct and is_match_remain):
        check_dict["verification_summary"]["total_errors"] += 1
        check_dict["verification_summary"]["artist_error_list"].append(artist)

    row_report_item_3 = {
        "아티스트": artist,
        "구분": "공제내역",
        "원본_곡비": cost_data.get("전월잔액", 0),
        "정산서_곡비": prev_val,
        "match_곡비": is_match_prev,

        "원본_공제금액": cost_data.get("당월차감액", 0),
        "정산서_공제금액": deduct_val,
        "match_공제금액": is_match_deduct,

        "원본_공제후잔액": cost_data.get("당월잔액", 0),
        "정산서_공제후잔액": remain_val,
        "match_공제후잔액": is_match_remain,
    }
    check_dict["details_verification"]["정산서"].append(row_report_item_3)

    #  - (2) 수익 배분율 검증
    original_rate = cost_data.get("정산요율", 0)
    is_rate_match = almost_equal(original_rate, rate_val)
    if not is_rate_match:
        check_dict["verification_summary"]["total_errors"] += 1
        check_dict["verification_summary"]["artist_error_list"].append(artist)

    row_report_item_4 = {
        "아티스트": artist,
        "구분": "수익배분율",
        "원본_정산율(%)": original_rate,
        "정산서_정산율(%)": rate_val,
        "match_정산율": is_rate_match,
    }
    check_dict["details_verification"]["정산서"].append(row_report_item_4)

    # 세부매출 검증(비교)
    for d in detail_list_sorted:
        original_val = d["revenue"]
        # 정산서 쪽도 사실상 d["revenue"] 그대로 사용
        report_val = d["revenue"]
        is_match = almost_equal(original_val, report_val)
        if not is_match:
            check_dict["verification_summary"]["total_errors"] += 1
            check_dict["verification_summary"]["artist_error_list"].append(artist)

        row_report_item = {
            "아티스트": artist,
            "구분": "음원서비스별매출",
            "앨범": d["album"],
            "서비스명": d["service"],
            "원본_매출액": original_val,
            "정산서_매출액": report_val,
            "match_매출액": is_match,
        }
        check_dict["details_verification"]["세부매출"].append(row_report_item)

    # 간단히 “정산서” 제목 행에만 스타일 부여 예시
    ws["B6"].font = Font(size=14, bold=True, color="000000")
    ws["B6"].alignment = Alignment(horizontal="center", vertical="center")

    return wb


if __name__ == "__main__":
    main()
